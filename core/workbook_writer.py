from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

from config.constants import (
    GENERATED_TFO_MANPOWER_MARKER,
    GENERATED_TFO_MARKER,
    GENERATED_TFO_RESULT_MARKER,
    MASTER_COLUMNS,
    MASTER_SHEET_NAME,
    TFO_INPUT_COLUMNS,
    TFO_MANPOWER_DISPLAY_COLUMNS,
    TFO_PRODUCTION_DISPLAY_COLUMNS,
    TFO_SHEET_NAME,
)
from core.dataframe_utils import clean_text, safe_float

TITLE_FILL = PatternFill("solid", fgColor="DCEBFF")
HEADER_FILL = PatternFill("solid", fgColor="EAF3FF")
BOLD_FONT = Font(bold=True)


def build_workbook_bytes(
    workbook_path: Path,
    master_df: pd.DataFrame,
    tfo_input_df: pd.DataFrame,
    tfo_production_df: pd.DataFrame,
    tfo_manpower_df: pd.DataFrame,
) -> bytes:
    workbook = openpyxl.load_workbook(workbook_path)
    master_sheet = workbook[MASTER_SHEET_NAME]
    tfo_sheet = workbook[TFO_SHEET_NAME]

    write_master_sheet(master_sheet=master_sheet, master_df=master_df)
    rewrite_tfo_sheet(
        tfo_sheet=tfo_sheet,
        tfo_input_df=tfo_input_df,
        tfo_production_df=tfo_production_df,
        tfo_manpower_df=tfo_manpower_df,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()



def write_state_to_workbook(
    source_workbook_path: Path,
    output_workbook_path: Path,
    master_df: pd.DataFrame,
    tfo_input_df: pd.DataFrame,
    tfo_production_df: pd.DataFrame,
    tfo_manpower_df: pd.DataFrame,
) -> Path:
    workbook = openpyxl.load_workbook(source_workbook_path)
    master_sheet = workbook[MASTER_SHEET_NAME]
    tfo_sheet = workbook[TFO_SHEET_NAME]

    write_master_sheet(master_sheet=master_sheet, master_df=master_df)
    rewrite_tfo_sheet(
        tfo_sheet=tfo_sheet,
        tfo_input_df=tfo_input_df,
        tfo_production_df=tfo_production_df,
        tfo_manpower_df=tfo_manpower_df,
    )

    output_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_workbook_path)
    return output_workbook_path



def write_master_sheet(master_sheet, master_df: pd.DataFrame) -> None:
    sorted_df = master_df.sort_values("__excel_row").reset_index(drop=True)

    for _, row in sorted_df.iterrows():
        excel_row = int(row["__excel_row"])

        for column_index, column_name in enumerate(MASTER_COLUMNS, start=1):
            value = row[column_name]
            if pd.isna(value):
                value = None
            master_sheet.cell(row=excel_row, column=column_index, value=value)



def _clear_sheet(sheet) -> None:
    if sheet.max_row > 0:
        sheet.delete_rows(1, sheet.max_row)



def _write_section_title(sheet, row_number: int, title_text: str, number_of_columns: int) -> None:
    sheet.cell(row=row_number, column=1, value=title_text)
    sheet.cell(row=row_number, column=1).font = BOLD_FONT
    sheet.cell(row=row_number, column=1).fill = TITLE_FILL
    for column_index in range(2, number_of_columns + 1):
        sheet.cell(row=row_number, column=column_index).fill = TITLE_FILL



def _write_table(sheet, start_row: int, dataframe: pd.DataFrame) -> int:
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = sheet.cell(row=start_row, column=column_index, value=column_name)
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL

    current_row = start_row + 1
    for _, row in dataframe.iterrows():
        for column_index, column_name in enumerate(dataframe.columns, start=1):
            value = row[column_name]
            if pd.isna(value):
                value = None
            sheet.cell(row=current_row, column=column_index, value=value)
        current_row += 1

    return current_row



def _autosize_columns(sheet, max_column: int) -> None:
    for column_index in range(1, max_column + 1):
        max_length = 12
        column_letter = openpyxl.utils.get_column_letter(column_index)
        for cell in sheet[column_letter]:
            text_value = clean_text(cell.value)
            if text_value:
                max_length = max(max_length, len(text_value) + 2)
        sheet.column_dimensions[column_letter].width = min(max_length, 34)



def rewrite_tfo_sheet(
    tfo_sheet,
    tfo_input_df: pd.DataFrame,
    tfo_production_df: pd.DataFrame,
    tfo_manpower_df: pd.DataFrame,
) -> None:
    _clear_sheet(tfo_sheet)

    input_df = tfo_input_df[TFO_INPUT_COLUMNS].copy()

    production_formula_columns = [
        "Production Required / Month Kgs Formula",
        "Production per Drum/day Formula",
        "No. of Drums Required Formula",
        "TFO Reqd./ Shift Formula",
        "kgs/drum/day Formula",
        "No. of Drums (A/W) Reqd. Formula",
        "Assembly Winding Reqd./ Shift Formula",
    ]
    production_df = tfo_production_df[TFO_PRODUCTION_DISPLAY_COLUMNS + production_formula_columns].copy()

    manpower_df = tfo_manpower_df[TFO_MANPOWER_DISPLAY_COLUMNS + ["Formulas"]].copy()

    _write_section_title(tfo_sheet, 1, GENERATED_TFO_MARKER, len(input_df.columns))
    next_row = _write_table(tfo_sheet, 2, input_df)

    production_title_row = next_row + 1
    _write_section_title(tfo_sheet, production_title_row, GENERATED_TFO_RESULT_MARKER, len(production_df.columns))
    next_row = _write_table(tfo_sheet, production_title_row + 1, production_df)

    manpower_title_row = next_row + 1
    _write_section_title(tfo_sheet, manpower_title_row, GENERATED_TFO_MANPOWER_MARKER, len(manpower_df.columns))
    _write_table(tfo_sheet, manpower_title_row + 1, manpower_df)

    _autosize_columns(tfo_sheet, max(len(input_df.columns), len(production_df.columns), len(manpower_df.columns)))
