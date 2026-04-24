from __future__ import annotations

from pathlib import Path
import shutil

import openpyxl
import pandas as pd

from config.constants import (
    DEFAULT_MACHINE_DIVISOR,
    DEFAULT_TFO_DIVISOR,
    GENERATED_TFO_MARKER,
    INPUT_WORKBOOK_CANDIDATES,
    MASTER_COLUMNS,
    MASTER_SHEET_NAME,
    OUTPUT_WORKBOOK_CANDIDATES,
    SPECIAL_TFO_COUNT,
    SPECIAL_TFO_DIVISOR,
    TFO_INPUT_COLUMNS,
    TFO_INPUT_END_ROW,
    TFO_INPUT_START_ROW,
    TFO_SHEET_NAME,
)
from core.dataframe_utils import clean_text, coerce_master_dataframe, normalize_key_text, safe_float


def get_input_workbook_path() -> Path:
    for candidate_path in INPUT_WORKBOOK_CANDIDATES:
        if candidate_path.exists():
            return candidate_path
    raise FileNotFoundError("input/Spinning.xlsx not found.")



def get_output_workbook_path() -> Path:
    for candidate_path in OUTPUT_WORKBOOK_CANDIDATES:
        if candidate_path.exists():
            return candidate_path
    return OUTPUT_WORKBOOK_CANDIDATES[0]



def resolve_workbook_path() -> Path:
    output_workbook_path = get_output_workbook_path()
    if output_workbook_path.exists():
        return output_workbook_path

    input_workbook_path = get_input_workbook_path()
    output_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_workbook_path, output_workbook_path)
    return output_workbook_path



def load_master_dataframe(workbook_path: Path) -> tuple[pd.DataFrame, list[str]]:
    master_df = pd.read_excel(workbook_path, sheet_name=MASTER_SHEET_NAME)

    for column_name in MASTER_COLUMNS:
        if column_name not in master_df.columns:
            master_df[column_name] = ""

    master_df = master_df[MASTER_COLUMNS].copy()
    master_df["__excel_row"] = range(2, len(master_df) + 2)
    master_df["__row_key"] = master_df["__excel_row"].apply(lambda row_number: f"ROW_{row_number}")

    section_order: list[str] = []
    seen_sections: set[str] = set()

    for section_name in master_df["Section"].tolist():
        section_text = clean_text(section_name)
        if not section_text:
            continue
        normalized_section = normalize_key_text(section_text)
        if normalized_section in seen_sections:
            continue
        seen_sections.add(normalized_section)
        section_order.append(section_text)

    return coerce_master_dataframe(master_df), section_order



def _default_tfo_divisor(count_value: object) -> float:
    if normalize_key_text(count_value) == normalize_key_text(SPECIAL_TFO_COUNT):
        return SPECIAL_TFO_DIVISOR
    return DEFAULT_TFO_DIVISOR



def _load_generated_tfo_input(sheet) -> pd.DataFrame:
    headers = [clean_text(sheet.cell(row=2, column=index).value) for index in range(1, len(TFO_INPUT_COLUMNS) + 1)]
    header_map = {index + 1: header for index, header in enumerate(headers) if header}

    rows = []
    current_row = 3
    while True:
        first_value = clean_text(sheet.cell(row=current_row, column=1).value)
        if not first_value:
            break

        row_dict = {}
        for column_index, header_name in header_map.items():
            row_dict[header_name] = sheet.cell(row=current_row, column=column_index).value

        normalized_row = {
            "Count": clean_text(row_dict.get("Count")),
            "Customer": clean_text(row_dict.get("Customer")),
            "Count2": safe_float(row_dict.get("Count2")),
            "Speed": safe_float(row_dict.get("Speed")),
            "TPI": safe_float(row_dict.get("TPI")),
            "Utilization": safe_float(row_dict.get("Utilization")),
            "Efficiency": safe_float(row_dict.get("Efficiency")),
            "Production Required / day Kgs": safe_float(row_dict.get("Production Required / day Kgs")),
            "TFO Divisor": safe_float(row_dict.get("TFO Divisor"), _default_tfo_divisor(row_dict.get("Count"))),
            "mpm": safe_float(row_dict.get("mpm")),
            "Eff": safe_float(row_dict.get("Eff")),
            "Machine Divisor": safe_float(row_dict.get("Machine Divisor"), DEFAULT_MACHINE_DIVISOR),
            "__excel_row": current_row,
        }
        rows.append(normalized_row)
        current_row += 1

    if not rows:
        return pd.DataFrame(columns=TFO_INPUT_COLUMNS + ["__excel_row"])

    generated_df = pd.DataFrame(rows)
    return generated_df[TFO_INPUT_COLUMNS + ["__excel_row"]].copy()



def _load_legacy_tfo_input(sheet) -> pd.DataFrame:
    rows = []
    for excel_row in range(TFO_INPUT_START_ROW, TFO_INPUT_END_ROW + 1):
        count_value = sheet[f"A{excel_row}"].value
        customer_value = sheet[f"B{excel_row}"].value

        rows.append(
            {
                "Count": clean_text(count_value),
                "Customer": clean_text(customer_value),
                "Count2": safe_float(sheet[f"C{excel_row}"].value),
                "Speed": safe_float(sheet[f"D{excel_row}"].value),
                "TPI": safe_float(sheet[f"E{excel_row}"].value),
                "Utilization": safe_float(sheet[f"F{excel_row}"].value),
                "Efficiency": safe_float(sheet[f"G{excel_row}"].value),
                "Production Required / day Kgs": safe_float(sheet[f"L{excel_row}"].value),
                "TFO Divisor": _default_tfo_divisor(count_value),
                "mpm": safe_float(sheet[f"Q{excel_row}"].value),
                "Eff": safe_float(sheet[f"R{excel_row}"].value),
                "Machine Divisor": DEFAULT_MACHINE_DIVISOR,
                "__excel_row": excel_row,
            }
        )

    legacy_df = pd.DataFrame(rows)
    return legacy_df[TFO_INPUT_COLUMNS + ["__excel_row"]].copy()



def load_tfo_input_dataframe(workbook_path: Path) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    sheet = workbook[TFO_SHEET_NAME]

    if clean_text(sheet["A1"].value) == GENERATED_TFO_MARKER:
        return _load_generated_tfo_input(sheet)

    return _load_legacy_tfo_input(sheet)
